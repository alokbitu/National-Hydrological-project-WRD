import csv
import os
from datetime import datetime
import openpyxl
import pandas as pd
import serial
import time
from config import *


def ser_com():
    with open ('raw_data.txt','r') as file :
        received_buffer = file.read()
        file.close()
    response = list(received_buffer.split('+'))
    return response

def calculation():
   received_buffer = ser_com()
   print(received_buffer)

   #calculating WATER LEVEL
   WL = received_buffer[1]
   print(WL,water_height,rl_value)
   water_level =  round(((((float(WL)- MinMa)/ (MaxMa-MinMa)) * water_height) + rl_value),3)
   print(water_level)


   #hourly & Daily rainfall calculation
   with open('current_rainfall.txt', 'r') as file:
       hourly_rainfall = file.read().strip()
       file.close()

   # Append the initial hourly rainfall to the lists
   hourly_rainfall_list.append(float(hourly_rainfall))
   daily_rainfall_list.append(float(hourly_rainfall))

   # Open the file again in 'w+' mode to write the updated value
   with open('current_rainfall.txt', 'w+') as file:
       # Set the initial value to '0.0'
       file.write('0.0')


   # to calculate the hourly rainfall
   with open('hourly_rainfall.txt','w+') as file:
       file.write(str(sum(hourly_rainfall_list)))
       file.close()
   if len(hourly_rainfall_list) <= 4:
       with open('hourly_rainfall.txt','r') as file2:
           hourly_rainfall = file2.read()
           print(hourly_rainfall)
           file2.close()
   if len(hourly_rainfall_list) >= 4:
       hourly_rainfall_list.clear()


   # to calculate the hourly rainfall
   with open('daily_rainfall.txt', 'w+') as file3:
       file3.write(str(sum(daily_rainfall_list)))
       file3.close()
   if len(daily_rainfall_list) <= 96:
       with open('daily_rainfall.txt','r') as file4:
           daily_rainfall = file4.read()
           print(daily_rainfall)
           file4.close()

   if len(daily_rainfall_list) >= 96:
       daily_rainfall_list.clear()

   raw_batt_voltage = received_buffer[5]
   diff_voltage = 12.566 - float(raw_batt_voltage)
   batt_voltage = float(raw_batt_voltage) + diff_voltage
   print(batt_voltage)
   solar_voltage = received_buffer[6]
   print(solar_voltage)

   date = datetime.now().strftime("%d/%m/%Y")
   Time = datetime.now().strftime("%H:%M")
   print(date)
   return water_level,hourly_rainfall,daily_rainfall,batt_voltage,solar_voltage,date,Time

def excel_create():
    water_level, hourly_rainfall, daily_rainfall, batt_voltage, solar_voltage, date, Time = calculation()

    # Function to check if the file is a valid Excel file
    def is_valid_excel(file_path):
        try:
            openpyxl.load_workbook(file_path)
            return True
        except openpyxl.utils.exceptions.InvalidFileException:
            return False

    # File name
    excel_file = 'data.xlsx'

    try:
        # Check if the file exists and is a valid Excel file
        if os.path.exists(excel_file) and is_valid_excel(excel_file):
            # Load the existing Excel workbook
            workbook = openpyxl.load_workbook(excel_file)
        else:
            # Delete the existing file if it's corrupted
            if os.path.exists(excel_file):
                os.remove(excel_file)
                print('curropt file get deleted')

            # Create a new Excel workbook
            workbook = openpyxl.Workbook()

        # Select the specific worksheet
        worksheet = workbook.active  # Use the active sheet in the workbook

        # Update the values in each row
        data = [
            ['Date', date],
            ['Time', Time],
            ['Solar Voltage', solar_voltage],
            ['Battery Voltage', batt_voltage],
            ['Water Depth', water_level],
            ['Hourly Rainfall', hourly_rainfall],
            ['Daily Rainfall', daily_rainfall],
        ]

        # Iterate over the data and update the worksheet
        for row_index, row_data in enumerate(data, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                worksheet.cell(row=row_index, column=col_index, value=cell_value)

        # Save the updated workbook
        workbook.save(excel_file)

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Close the workbook to release resources
        if 'workbook' in locals():
            workbook.close()
            print('workbook updated')
        return water_level,hourly_rainfall,daily_rainfall,batt_voltage,solar_voltage,date,Time


def ftp_csv():
    water_level,hourly_rainfall,daily_rainfall,batt_volatage,solar_voltage,date,Time = excel_create()
    # Create the file name in the specified format
    current_datetime = datetime.now()
    date_part = current_datetime.strftime("%y%m%d")
    time_part = current_datetime.strftime("%H%M%S")
    file_name = f"{location}_{date_part}_{time_part}_{station_number}.csv"

    # Create the file path
    file_path = "E://sunjray job documents//projects//National-Hydrological-project-WRD//Backend//WRD_Received//" + file_name
    file_path2 = "E://sunjray job documents//projects//National-Hydrological-project-WRD//Backend//Sunjray_Received//"+file_name
    print(date)
    # Create a list of values to be written in a single row
    values = [
        station_id,
        date+' '+Time,
        sim_number,
        batt_volatage,
        water_level,
        hourly_rainfall,  # Hourly Rainfall
        daily_rainfall,  # Daily Rainfall
        solar_voltage
    ]

    # Create a new CSV file and write data to it
    with open(file_path, "w", newline="") as file:
        csv_writer = csv.writer(file)
        csv_writer.writerow(values)
    with open(file_path2, "w", newline="") as file2:
        csv_writer = csv.writer(file2)
        csv_writer.writerow(values)

    print(f"CSV file '{file_name}' created successfully.")

if __name__  == "__main__":
     while True:
        try:
            ftp_csv()
        except Exception as e:
            print("An error occurred:", e)
            print("Restarting the code...")
            time.sleep(10)  # Delay before rerunning the code
            continue
        else:
            print("Code executed successfully.")
        finally:
            time.sleep(60)  # Delay between each iteration
